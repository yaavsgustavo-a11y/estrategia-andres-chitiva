import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from datetime import datetime, date
from collections import Counter
import re
import os

# ── Paths ──────────────────────────────────────────────────────────────────────
SRC = "Copia de BASE DE DATOS LEADS CHITIVA JULIO .xlsx"
DST = "DASHBOARD_COMERCIAL_CHITIVA.xlsx"

# ── Palette ───────────────────────────────────────────────────────────────────
C_DARK   = "1A1A2E"
C_MID    = "16213E"
C_ACC1   = "0F3460"
C_ACC2   = "E94560"
C_GREEN  = "00B894"
C_YELLOW = "FDCB6E"
C_ORANGE = "E17055"
C_WHITE  = "FFFFFF"
C_LGRAY  = "F0F2F5"
C_GRAY   = "BDC3C7"
C_TEXT   = "2D3436"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=11, color=C_TEXT, italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_all(color=C_GRAY):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom(color=C_ACC1):
    s = Side(style="medium", color=color)
    return Border(bottom=s)

# ── Source data load ──────────────────────────────────────────────────────────
wb_src = load_workbook(SRC, data_only=True)
ws_src = wb_src["Hoja1"]
print(f"Fuente cargada: {ws_src.max_row} filas x {ws_src.max_column} cols")


# ── Source columns (0-indexed) ────────────────────────────────────────────────
# Row 2 = headers: A=FECHA, B=Nombre, C=Telefono, D=Correo, E=Interes,
#                  F=Medio, G=Anuncio, H=Canal, I=Etapa, J=Resultado, K=Comentario

FUENTE_MAP = {
    "FACEBOOK": "Facebook",
    "INSTAGRAM": "Instagram",
    "TIK TOK": "TikTok",
    "TIKTOK": "TikTok",
    "TIKtok": "TikTok",
    "CONOCIDO": "Referido/Conocido",
    "WHATSAPP": "WhatsApp",
    "CATELOGO DE WHATSAPP": "WhatsApp",
    "CATALOGO DE WHATSAPP": "WhatsApp",
    "CERCANIA DE INSTALACIONES": "Cercanía",
    "CERCANIA": "Cercanía",
    "VIDEO": "Orgánico",
}

def norm_fuente(val):
    if not val or str(val).strip() in ("-", ""):
        return "Sin datos"
    v = str(val).strip().upper()
    for k, mapped in FUENTE_MAP.items():
        if v == k.upper():
            return mapped
    if "FACEBOOK" in v:
        return "Facebook"
    if "INSTAGRAM" in v:
        return "Instagram"
    if "TIKTOK" in v or "TIK" in v:
        return "TikTok"
    if "WHATSAPP" in v or "CATAL" in v:
        return "WhatsApp"
    if "CONOCIDO" in v or "REFERIDO" in v:
        return "Referido/Conocido"
    if "CERCANIA" in v or "CERCANÍA" in v:
        return "Cercanía"
    return "Otro"

ESTATUS_ORDER = {
    "6 - Inscrito": 6,
    "5 - En proceso inscripción": 5,
    "4 - Clase muestra": 4,
    "3 - Contactado/Activo": 3,
    "2 - Interesado": 2,
    "7 - Perdido": 7,
    "8 - Sin respuesta": 8,
    "1 - Lead/Informes": 1,
}

def norm_estatus(val):
    if not val or str(val).strip() in ("-", ""):
        return "1 - Lead/Informes"
    v = str(val).strip().upper()
    if v == "INSCRITO":
        return "6 - Inscrito"
    if "PROCESO DE INSCRIPCI" in v or v == "PROCESO DE INSCRIPCION":
        return "5 - En proceso inscripción"
    if "CLASE MUESTRA" in v:
        return "4 - Clase muestra"
    if any(x in v for x in ["PARTIDO AMISTOSO","TRATO","CONVENIO","RENTA"]):
        return "3 - Contactado/Activo"
    if any(x in v for x in ["INTERESADO","ESPERANDO","EN ESPERA","ESPERA","ELECCIÓN","ELECCION"]):
        return "2 - Interesado"
    if any(x in v for x in ["DESINTERESADO","NO INSCRITO","SIN LIGA"]):
        return "7 - Perdido"
    if any(x in v for x in ["SIN RESPUESTA","SIN CONTESTACI","DEJO DE CONTESTAR",
                              "NO RESPONDIO","NO VOLVIO","NO CONTESTO","NO CONTESTA"]):
        return "8 - Sin respuesta"
    return "1 - Lead/Informes"

def get_week(d):
    if isinstance(d, (datetime, date)):
        if isinstance(d, datetime):
            d = d.date()
        d2 = date(d.year, d.month, d.day)
        wd = d2.isocalendar()
        return f"{wd[0]}-W{wd[1]:02d}"
    return None

def get_month(d):
    if isinstance(d, (datetime, date)):
        if isinstance(d, datetime):
            d = d.date()
        return d.strftime("%Y-%m")
    return None


# ── Extract & clean rows ──────────────────────────────────────────────────────
leads = []
for row in ws_src.iter_rows(min_row=3, values_only=True):
    fecha_raw  = row[0]
    nombre     = row[1]
    telefono   = row[2]
    correo     = row[3]
    interes    = row[4]
    medio      = row[5]
    anuncio    = row[6]
    canal      = row[7]
    etapa      = row[8]
    resultado  = row[9]
    comentario = row[10]

    # Skip rows without a real name
    if not nombre or str(nombre).strip() in ("-", "", "None"):
        continue

    # Fix 2028 → 2026 (known data-entry error)
    if isinstance(fecha_raw, datetime) and fecha_raw.year == 2028:
        fecha_raw = fecha_raw.replace(year=2026)
    elif isinstance(fecha_raw, datetime) and fecha_raw.year == 2027:
        fecha_raw = fecha_raw.replace(year=2026)

    fecha_date = fecha_raw.date() if isinstance(fecha_raw, datetime) else fecha_raw

    fuente = norm_fuente(anuncio if anuncio and str(anuncio).strip() not in ("-","") else medio)
    estatus = norm_estatus(resultado)
    semana  = get_week(fecha_date)
    mes     = get_month(fecha_date)

    leads.append({
        "fecha":     fecha_date,
        "nombre":    str(nombre).strip(),
        "telefono":  str(telefono).replace(".0","") if telefono else "",
        "correo":    str(correo).strip() if correo and str(correo)!="-" else "",
        "interes":   str(interes).strip().title() if interes and str(interes)!="-" else "",
        "medio":     str(medio).strip().title() if medio and str(medio)!="-" else "",
        "anuncio":   str(anuncio).strip().title() if anuncio and str(anuncio)!="-" else "",
        "canal":     "WhatsApp",
        "etapa":     str(etapa).strip().title() if etapa and str(etapa)!="-" else "",
        "resultado": str(resultado).strip().title() if resultado and str(resultado)!="-" else "",
        "comentario":str(comentario).strip() if comentario and str(comentario)!="-" else "",
        "fuente":    fuente,
        "estatus":   estatus,
        "semana":    semana,
        "mes":       mes,
    })

print(f"Leads limpios: {len(leads)}")
total = len(leads)
inscritos   = sum(1 for l in leads if l["estatus"] == "6 - Inscrito")
en_proceso  = sum(1 for l in leads if l["estatus"] == "5 - En proceso inscripción")
interesados = sum(1 for l in leads if l["estatus"] == "2 - Interesado")
activos     = sum(1 for l in leads if l["estatus"] == "3 - Contactado/Activo")
perdidos    = sum(1 for l in leads if l["estatus"] == "7 - Perdido")
sin_resp    = sum(1 for l in leads if l["estatus"] == "8 - Sin respuesta")
tasa_conv   = inscritos / total if total else 0
print(f"Total:{total}  Inscritos:{inscritos}  Perdidos:{perdidos}  Tasa:{tasa_conv:.1%}")

# Aggregations
by_fuente = Counter(l["fuente"] for l in leads)
ins_by_fuente = Counter(l["fuente"] for l in leads if l["estatus"]=="6 - Inscrito")
by_mes = Counter(l["mes"] for l in leads if l["mes"])
ins_by_mes = Counter(l["mes"] for l in leads if l["mes"] and l["estatus"]=="6 - Inscrito")
by_estatus = Counter(l["estatus"] for l in leads)
by_semana = Counter(l["semana"] for l in leads if l["semana"])
ins_by_semana = Counter(l["semana"] for l in leads if l["semana"] and l["estatus"]=="6 - Inscrito")
by_interes = Counter(l["interes"] for l in leads if l["interes"])
meses_sorted = sorted(by_mes.keys())
semanas_sorted = sorted(by_semana.keys())


# ── Build new workbook ────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════════════════════
# HOJA 1 — LEADS_DB (cleaned)
# ══════════════════════════════════════════════════════════════════════════════
ws_db = wb.create_sheet("LEADS_DB")
ws_db.sheet_properties.tabColor = "0F3460"

DB_HEADERS = [
    "ID", "FECHA", "NOMBRE", "TELÉFONO", "CORREO",
    "INTERÉS", "MEDIO CONTACTO", "ANUNCIO/FUENTE", "CANAL SEGUIMIENTO",
    "ETAPA PROCESO", "RESULTADO", "COMENTARIO",
    "FUENTE LIMPIA", "ESTATUS FINAL", "SEMANA", "MES"
]

# Freeze panes & dimensions
ws_db.freeze_panes = "A3"
col_widths_db = [5,13,26,16,24,22,18,22,20,22,22,28,20,26,12,10]
for i, w in enumerate(col_widths_db, 1):
    ws_db.column_dimensions[get_column_letter(i)].width = w

# Title row
ws_db.row_dimensions[1].height = 30
ws_db.merge_cells("A1:P1")
tc = ws_db["A1"]
tc.value = "📋  BASE DE DATOS — LEADS COMERCIALES  |  Andres Chitiva"
tc.font = font(bold=True, size=13, color=C_WHITE)
tc.fill = fill(C_DARK)
tc.alignment = align()

# Header row
ws_db.row_dimensions[2].height = 22
for ci, h in enumerate(DB_HEADERS, 1):
    cell = ws_db.cell(row=2, column=ci, value=h)
    cell.font = font(bold=True, size=9, color=C_WHITE)
    cell.fill = fill(C_ACC1)
    cell.alignment = align(wrap=True)
    cell.border = border_all(C_ACC1)

# Data rows
ROW_FILLS = [fill(C_LGRAY), fill(C_WHITE)]
for i, lead in enumerate(leads, 1):
    r = i + 2
    row_fill = ROW_FILLS[i % 2]
    ws_db.row_dimensions[r].height = 16
    values = [
        i,
        lead["fecha"],
        lead["nombre"],
        lead["telefono"],
        lead["correo"],
        lead["interes"],
        lead["medio"],
        lead["anuncio"],
        lead["canal"],
        lead["etapa"],
        lead["resultado"],
        lead["comentario"],
        lead["fuente"],
        lead["estatus"],
        lead["semana"],
        lead["mes"],
    ]
    for ci, v in enumerate(values, 1):
        cell = ws_db.cell(row=r, column=ci, value=v)
        cell.font = font(size=9)
        cell.fill = row_fill
        cell.border = border_all()
        cell.alignment = align(h="left" if ci > 2 else "center", wrap=False)
    # Color estatus cell
    est_cell = ws_db.cell(row=r, column=14)
    est = lead["estatus"]
    ec = {"6 - Inscrito": C_GREEN, "7 - Perdido": C_ACC2,
          "8 - Sin respuesta": C_ORANGE, "2 - Interesado": C_YELLOW,
          "5 - En proceso inscripción": "74B9FF"}.get(est)
    if ec:
        est_cell.fill = fill(ec)
        est_cell.font = font(size=9, bold=True,
                             color=C_WHITE if ec not in (C_YELLOW,) else C_TEXT)

print("LEADS_DB creada")


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 2 — CONFIG
# ══════════════════════════════════════════════════════════════════════════════
ws_cfg = wb.create_sheet("Config")
ws_cfg.sheet_properties.tabColor = "636E72"
ws_cfg.column_dimensions["A"].width = 30
ws_cfg.column_dimensions["C"].width = 25
ws_cfg.column_dimensions["E"].width = 22

def write_config_block(ws, col, title, items, title_fill=C_ACC1):
    ws.cell(row=1, column=col, value=title).font = font(bold=True, size=10, color=C_WHITE)
    ws.cell(row=1, column=col).fill = fill(title_fill)
    ws.cell(row=1, column=col).alignment = align()
    for i, item in enumerate(items, 2):
        c = ws.cell(row=i, column=col, value=item)
        c.font = font(size=9)
        c.fill = fill(C_LGRAY if i % 2 == 0 else C_WHITE)
        c.border = border_all()

ESTATUS_LIST = [
    "1 - Lead/Informes",
    "2 - Interesado",
    "3 - Contactado/Activo",
    "4 - Clase muestra",
    "5 - En proceso inscripción",
    "6 - Inscrito",
    "7 - Perdido",
    "8 - Sin respuesta",
]

FUENTES_LIST = ["Facebook","Instagram","TikTok","WhatsApp",
                "Referido/Conocido","Cercanía","Orgánico","Otro","Sin datos"]

INTERESES_LIST = ["Torneo","Curso de Verano","Curso Piloto","Pretemporada Intensiva",
                  "Academia","Academia de Fútbol","Empresarial","Liga Infantil",
                  "Veterano","Fin de Semana","Segunda Varonil","Tercera Varonil",
                  "Premier Varonil","General","Otro"]

write_config_block(ws_cfg, 1, "ESTATUS FINAL (Embudo)", ESTATUS_LIST, C_ACC1)
write_config_block(ws_cfg, 3, "FUENTES", FUENTES_LIST, C_MID)
write_config_block(ws_cfg, 5, "INTERESES", INTERESES_LIST, C_ACC2)
print("Config creada")

# ══════════════════════════════════════════════════════════════════════════════
# HOJA 3 — DATOS_PIVOT
# ══════════════════════════════════════════════════════════════════════════════
ws_pv = wb.create_sheet("DATOS_PIVOT")
ws_pv.sheet_properties.tabColor = "0984E3"

def pv_header(ws, row, col, text, bg=C_ACC1):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=True, size=10, color=C_WHITE)
    c.fill = fill(bg)
    c.alignment = align()
    c.border = border_all(bg)
    return c

def pv_section_title(ws, row, col, text, span=4):
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=True, size=11, color=C_WHITE)
    c.fill = fill(C_DARK)
    c.alignment = align()

def pv_val(ws, row, col, value, fmt=None, bg=C_WHITE):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(size=10)
    c.fill = fill(bg)
    c.border = border_all()
    c.alignment = align()
    if fmt:
        c.number_format = fmt
    return c

# ── Section 1: KPIs Generales (col A) ─────────────────────────────────────
pv_section_title(ws_pv, 1, 1, "📊  KPIs GENERALES", 3)
pv_header(ws_pv, 2, 1, "MÉTRICA")
pv_header(ws_pv, 2, 2, "VALOR")
pv_header(ws_pv, 2, 3, "DETALLE")
ws_pv.column_dimensions["A"].width = 26
ws_pv.column_dimensions["B"].width = 14
ws_pv.column_dimensions["C"].width = 22

kpis = [
    ("Total Leads",           total,        "Leads con nombre registrado"),
    ("Inscritos",             inscritos,     "Resultado = Inscrito"),
    ("En proceso inscripción",en_proceso,   "Casi convertidos"),
    ("Interesados activos",   interesados,  "Seguimiento activo"),
    ("Contactados/Activos",   activos,       "Partido amistoso / trato"),
    ("Perdidos",              perdidos,      "Desinteresados"),
    ("Sin respuesta",         sin_resp,      "No contestaron"),
    ("Tasa de Conversión",    tasa_conv,     f"{inscritos} de {total} leads"),
    ("Tasa de Pérdida",       perdidos/total if total else 0, f"{perdidos} de {total} leads"),
    ("Pendientes activos",    total-inscritos-perdidos-sin_resp, "Leads en proceso"),
]
for i, (label, val, det) in enumerate(kpis, 3):
    pv_val(ws_pv, i, 1, label, bg=C_LGRAY if i%2==0 else C_WHITE)
    fmt = "0.0%" if isinstance(val, float) else "#,##0"
    pv_val(ws_pv, i, 2, val, fmt=fmt,
           bg=C_GREEN if label=="Inscritos" else (C_ACC2 if label=="Perdidos" else C_WHITE))
    pv_val(ws_pv, i, 3, det, bg=C_LGRAY if i%2==0 else C_WHITE)
print("DATOS_PIVOT KPIs ok")


# ── Section 2: Embudo (col E) ──────────────────────────────────────────────
pv_section_title(ws_pv, 1, 5, "🔽  EMBUDO DE CONVERSIÓN", 4)
for c_idx, h in enumerate(["ETAPA","TOTAL","% DEL TOTAL","% CONV. ETAPA ANT."], 5):
    pv_header(ws_pv, 2, c_idx, h, C_MID)
ws_pv.column_dimensions["E"].width = 30
ws_pv.column_dimensions["F"].width = 10
ws_pv.column_dimensions["G"].width = 14
ws_pv.column_dimensions["H"].width = 18

# Embudo acumulativo
funnel = [
    ("Lead recibido",           total),
    ("Contactado / Informes",   total - sum(1 for l in leads if l["estatus"]=="8 - Sin respuesta")),
    ("Interesado",              inscritos + en_proceso + activos + interesados),
    ("Clase muestra",           inscritos + en_proceso + sum(1 for l in leads if l["estatus"]=="4 - Clase muestra")),
    ("En proceso inscripción",  inscritos + en_proceso),
    ("Inscrito ✅",              inscritos),
]
prev = total
for i, (label, val) in enumerate(funnel, 3):
    pv_val(ws_pv, i, 5, label, bg=C_LGRAY if i%2==0 else C_WHITE)
    pv_val(ws_pv, i, 6, val)
    pv_val(ws_pv, i, 7, val/total if total else 0, fmt="0.0%")
    pv_val(ws_pv, i, 8, val/prev if prev else 0, fmt="0.0%")
    prev = val if val > 0 else 1

# ── Section 3: Por Fuente (col J) ─────────────────────────────────────────
pv_section_title(ws_pv, 1, 10, "📡  LEADS & CONVERSIÓN POR FUENTE", 5)
for c_idx, h in enumerate(["FUENTE","LEADS","INSCRITOS","TASA CONV.","% DEL TOTAL"], 10):
    pv_header(ws_pv, 2, c_idx, h, C_ACC2)
ws_pv.column_dimensions["J"].width = 22
ws_pv.column_dimensions["K"].width = 10
ws_pv.column_dimensions["L"].width = 12
ws_pv.column_dimensions["M"].width = 13
ws_pv.column_dimensions["N"].width = 14

fuentes_sorted = sorted(by_fuente.items(), key=lambda x: -x[1])
for i, (fuente, cnt) in enumerate(fuentes_sorted, 3):
    ins = ins_by_fuente.get(fuente, 0)
    pv_val(ws_pv, i, 10, fuente, bg=C_LGRAY if i%2==0 else C_WHITE)
    pv_val(ws_pv, i, 11, cnt)
    pv_val(ws_pv, i, 12, ins,
           bg=fill(C_GREEN).fgColor.rgb if ins > 0 else C_WHITE)
    pv_val(ws_pv, i, 13, ins/cnt if cnt else 0, fmt="0.0%")
    pv_val(ws_pv, i, 14, cnt/total if total else 0, fmt="0.0%")
print("DATOS_PIVOT Fuente ok")

# ── Section 4: Por Mes (col A, below row 14) ──────────────────────────────
START_ROW_MES = 15
pv_section_title(ws_pv, START_ROW_MES, 1, "📅  EVOLUCIÓN MENSUAL", 4)
for c_idx, h in enumerate(["MES","LEADS","INSCRITOS","TASA CONV."], 1):
    pv_header(ws_pv, START_ROW_MES+1, c_idx, h, C_ACC1)

for i, mes in enumerate(meses_sorted, START_ROW_MES+2):
    cnt = by_mes[mes]
    ins = ins_by_mes.get(mes, 0)
    pv_val(ws_pv, i, 1, mes, bg=C_LGRAY if i%2==0 else C_WHITE)
    pv_val(ws_pv, i, 2, cnt)
    pv_val(ws_pv, i, 3, ins)
    pv_val(ws_pv, i, 4, ins/cnt if cnt else 0, fmt="0.0%")

# ── Section 5: Por Semana (col E, below row 14) ───────────────────────────
pv_section_title(ws_pv, START_ROW_MES, 5, "📆  EVOLUCIÓN SEMANAL", 4)
for c_idx, h in enumerate(["SEMANA","LEADS","INSCRITOS","TASA CONV."], 5):
    pv_header(ws_pv, START_ROW_MES+1, c_idx, h, C_MID)

for i, sem in enumerate(semanas_sorted, START_ROW_MES+2):
    cnt = by_semana[sem]
    ins = ins_by_semana.get(sem, 0)
    pv_val(ws_pv, i, 5, sem, bg=C_LGRAY if i%2==0 else C_WHITE)
    pv_val(ws_pv, i, 6, cnt)
    pv_val(ws_pv, i, 7, ins)
    pv_val(ws_pv, i, 8, ins/cnt if cnt else 0, fmt="0.0%")

# ── Section 6: Por Interés (col J, below row 14) ──────────────────────────
pv_section_title(ws_pv, START_ROW_MES, 10, "🎯  LEADS POR TIPO DE INTERÉS", 5)
for c_idx, h in enumerate(["INTERÉS","LEADS","INSCRITOS","TASA CONV.","% TOTAL"], 10):
    pv_header(ws_pv, START_ROW_MES+1, c_idx, h, C_ACC2)

ins_by_interes = Counter(l["interes"] for l in leads if l["interes"] and l["estatus"]=="6 - Inscrito")
interes_sorted = sorted(by_interes.items(), key=lambda x: -x[1])
for i, (interes, cnt) in enumerate(interes_sorted, START_ROW_MES+2):
    ins = ins_by_interes.get(interes, 0)
    pv_val(ws_pv, i, 10, interes, bg=C_LGRAY if i%2==0 else C_WHITE)
    pv_val(ws_pv, i, 11, cnt)
    pv_val(ws_pv, i, 12, ins)
    pv_val(ws_pv, i, 13, ins/cnt if cnt else 0, fmt="0.0%")
    pv_val(ws_pv, i, 14, cnt/total if total else 0, fmt="0.0%")

print("DATOS_PIVOT completo")


# ══════════════════════════════════════════════════════════════════════════════
# HOJA 4 — DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
ws_d = wb.create_sheet("DASHBOARD")
ws_d.sheet_properties.tabColor = "E94560"
ws_d.sheet_view.showGridLines = False
ws_d.freeze_panes = "A4"

# Column widths
for col_i, w in enumerate([1,14,14,14,14,14,14,14,2,16,16,16,16,16,2,14,14,14,14,14], 1):
    ws_d.column_dimensions[get_column_letter(col_i)].width = w
# Row heights
for r, h in [(1,45),(2,20),(3,14),(4,12),(5,50),(6,50),(7,14),
             (8,50),(9,50),(10,14),(11,50),(12,50),(13,14),(14,35)]:
    ws_d.row_dimensions[r].height = h

def dash_fill_range(ws, r1, c1, r2, c2, hex_color):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).fill = fill(hex_color)

# ── Background ────────────────────────────────────────────────────────────
dash_fill_range(ws_d, 1, 1, 80, 22, C_DARK)

# ── Main Title ────────────────────────────────────────────────────────────
ws_d.merge_cells("B1:H1")
t = ws_d["B1"]
t.value = "DASHBOARD COMERCIAL"
t.font = Font(name="Calibri", bold=True, size=22, color=C_WHITE)
t.alignment = align()

ws_d.merge_cells("J1:T1")
t2 = ws_d["J1"]
t2.value = "Andres Chitiva  ·  Julio 2026"
t2.font = Font(name="Calibri", bold=False, size=14, color=C_GRAY, italic=True)
t2.alignment = align(h="right")

# Subtitle line
ws_d.merge_cells("B2:T2")
sub = ws_d["B2"]
sub.value = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Fuente: LEADS_DB"
sub.font = Font(name="Calibri", size=9, color=C_GRAY, italic=True)
sub.alignment = align(h="left")

# ── KPI Cards (row 4-6) ───────────────────────────────────────────────────
cards = [
    ("B", "C",  C_GREEN,  "INSCRITOS",         inscritos,    f"de {total} leads"),
    ("D", "E",  C_ACC2,   "PERDIDOS",           perdidos,     "desinteresados"),
    ("F", "G",  C_YELLOW, "SIN RESPUESTA",       sin_resp,     "no contestaron"),
    ("J", "K",  "74B9FF", "EN PROCESO",         en_proceso,   "casi inscritos"),
    ("L", "M",  C_ORANGE, "INTERESADOS",        interesados,  "seguimiento activo"),
    ("N", "O",  "A29BFE", "TASA CONVERSIÓN",    f"{tasa_conv:.1%}", f"base {total} leads"),
]

for (c1l, c2l, bg, label, value, sub_text) in cards:
    c1 = openpyxl.utils.column_index_from_string(c1l)
    c2 = openpyxl.utils.column_index_from_string(c2l)
    ws_d.merge_cells(start_row=4, start_column=c1, end_row=4, end_column=c2)
    ws_d.merge_cells(start_row=5, start_column=c1, end_row=5, end_column=c2)
    ws_d.merge_cells(start_row=6, start_column=c1, end_row=6, end_column=c2)
    dash_fill_range(ws_d, 4, c1, 6, c2, bg)
    lbl = ws_d.cell(row=4, column=c1, value=label)
    lbl.font = Font(name="Calibri", bold=True, size=9, color=C_TEXT)
    lbl.alignment = align()
    val_c = ws_d.cell(row=5, column=c1, value=value)
    val_c.font = Font(name="Calibri", bold=True, size=24, color=C_TEXT)
    val_c.alignment = align()
    sub_c = ws_d.cell(row=6, column=c1, value=sub_text)
    sub_c.font = Font(name="Calibri", size=8, color=C_TEXT, italic=True)
    sub_c.alignment = align()

# Total Leads card
ws_d.merge_cells("B4:C4"); ws_d.merge_cells("B5:C5"); ws_d.merge_cells("B6:C6")
# overwrite with total
ws_d.cell(row=4, column=2, value="TOTAL LEADS").font = Font(name="Calibri",bold=True,size=9,color=C_WHITE)
ws_d.cell(row=4, column=2).fill = fill(C_ACC1)
ws_d.cell(row=4, column=2).alignment = align()
ws_d.cell(row=5, column=2, value=total).font = Font(name="Calibri",bold=True,size=28,color=C_WHITE)
ws_d.cell(row=5, column=2).fill = fill(C_ACC1)
ws_d.cell(row=5, column=2).alignment = align()
ws_d.cell(row=6, column=2, value="leads registrados").font = Font(name="Calibri",size=8,color=C_GRAY,italic=True)
ws_d.cell(row=6, column=2).fill = fill(C_ACC1)
ws_d.cell(row=6, column=2).alignment = align()
print("Dashboard KPI cards ok")


# ── Section labels ─────────────────────────────────────────────────────────
def dash_section_label(ws, row, col_start, col_end, text, bg=C_MID):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.font = Font(name="Calibri", bold=True, size=10, color=C_WHITE)
    c.fill = fill(bg)
    c.alignment = align(h="left")
    c.border = Border(left=Side(style="thick", color=C_ACC2),
                      bottom=Side(style="thin", color=C_GRAY))

def dash_table_header(ws, row, cols_vals, bg=C_ACC1):
    for ci, (col, val) in enumerate(cols_vals):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", bold=True, size=8, color=C_WHITE)
        c.fill = fill(bg)
        c.alignment = align()
        c.border = border_all(bg)

def dash_table_row(ws, row, col_val_list, alt=False):
    bg = C_LGRAY if alt else "ECEFF1"
    for col, val, fmt in col_val_list:
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", size=9, color=C_TEXT)
        c.fill = fill(bg)
        c.alignment = align()
        c.border = border_all()
        if fmt:
            c.number_format = fmt

# ── EMBUDO TABLE (rows 8-15, cols B-E) ────────────────────────────────────
dash_section_label(ws_d, 8, 2, 6, "🔽  EMBUDO DE CONVERSIÓN", C_ACC1)
dash_table_header(ws_d, 9, [(2,"ETAPA"),(4,"TOTAL"),(5,"% TOTAL"),(6,"% ETAPA ANT.")], C_MID)
ws_d.merge_cells("B9:C9")

for i, (label, val) in enumerate(funnel, 10):
    prev_val = funnel[i-11][1] if i > 10 else total
    ws_d.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
    dash_table_row(ws_d, i, [
        (2, label, None),
        (4, val, "#,##0"),
        (5, val/total if total else 0, "0.0%"),
        (6, val/prev_val if prev_val else 0, "0.0%"),
    ], alt=i%2==0)
    # Progress bar visual using cell background intensity
    pct = val / total if total else 0
    bar_color = C_GREEN if label.startswith("Inscrito") else "74B9FF"
    ws_d.cell(row=i, column=4).fill = fill(bar_color if pct > 0.5 else C_ACC1)
    ws_d.cell(row=i, column=4).font = Font(name="Calibri", bold=True, size=9, color=C_WHITE)

# ── FUENTE TABLE (rows 8-15, cols J-N) ────────────────────────────────────
dash_section_label(ws_d, 8, 10, 16, "📡  LEADS POR FUENTE Y CONVERSIÓN", C_ACC2)
dash_table_header(ws_d, 9, [(10,"FUENTE"),(12,"LEADS"),(13,"INSCRITOS"),(14,"TASA"),(15,"% TOTAL")], C_DARK)
ws_d.merge_cells("J9:K9")

for i, (fuente, cnt) in enumerate(fuentes_sorted, 10):
    ins = ins_by_fuente.get(fuente, 0)
    ws_d.merge_cells(start_row=i, start_column=10, end_row=i, end_column=11)
    dash_table_row(ws_d, i, [
        (10, fuente, None),
        (12, cnt, "#,##0"),
        (13, ins, "#,##0"),
        (14, ins/cnt if cnt else 0, "0.0%"),
        (15, cnt/total if total else 0, "0.0%"),
    ], alt=i%2==0)

print("Dashboard tables ok")


# ── EVOLUCIÓN MENSUAL TABLE (rows 18-24, cols B-G) ─────────────────────────
START_EVO = 18
dash_section_label(ws_d, START_EVO, 2, 7, "📅  EVOLUCIÓN MENSUAL", C_ACC1)
dash_table_header(ws_d, START_EVO+1, [(2,"MES"),(4,"LEADS"),(5,"INSCRITOS"),(6,"TASA"),(7,"VARIACIÓN")], C_MID)
ws_d.merge_cells(f"B{START_EVO+1}:C{START_EVO+1}")

prev_leads = None
for i, mes in enumerate(meses_sorted, START_EVO+2):
    cnt = by_mes[mes]
    ins = ins_by_mes.get(mes, 0)
    variacion = (cnt - prev_leads) / prev_leads if prev_leads else 0
    ws_d.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
    dash_table_row(ws_d, i, [
        (2, mes, None),
        (4, cnt, "#,##0"),
        (5, ins, "#,##0"),
        (6, ins/cnt if cnt else 0, "0.0%"),
        (7, variacion, "+0.0%;-0.0%;0.0%"),
    ], alt=i%2==0)
    var_c = ws_d.cell(row=i, column=7)
    var_c.font = Font(name="Calibri", size=9,
                      color=C_GREEN if variacion >= 0 else C_ACC2, bold=True)
    prev_leads = cnt

# ── EVOLUCIÓN SEMANAL TABLE (rows 18-xx, cols J-N) ────────────────────────
dash_section_label(ws_d, START_EVO, 10, 15, "📆  EVOLUCIÓN SEMANAL", C_ACC2)
dash_table_header(ws_d, START_EVO+1, [(10,"SEMANA"),(12,"LEADS"),(13,"INSCRITOS"),(14,"TASA")], C_DARK)
ws_d.merge_cells(f"J{START_EVO+1}:K{START_EVO+1}")

for i, sem in enumerate(semanas_sorted, START_EVO+2):
    cnt = by_semana[sem]
    ins = ins_by_semana.get(sem, 0)
    ws_d.merge_cells(start_row=i, start_column=10, end_row=i, end_column=11)
    dash_table_row(ws_d, i, [
        (10, sem, None),
        (12, cnt, "#,##0"),
        (13, ins, "#,##0"),
        (14, ins/cnt if cnt else 0, "0.0%"),
    ], alt=i%2==0)

# ── TOP INTERESES TABLE ────────────────────────────────────────────────────
START_INT = max(START_EVO + len(meses_sorted) + 4, START_EVO + len(semanas_sorted) + 4)
dash_section_label(ws_d, START_INT, 2, 8, "🎯  TOP INTERESES / PRODUCTOS", C_MID)
dash_table_header(ws_d, START_INT+1,
    [(2,"INTERÉS"),(5,"LEADS"),(6,"INSCRITOS"),(7,"TASA"),(8,"% TOTAL")], C_ACC1)
ws_d.merge_cells(f"B{START_INT+1}:D{START_INT+1}")

for i, (interes, cnt) in enumerate(interes_sorted[:12], START_INT+2):
    ins = ins_by_interes.get(interes, 0)
    ws_d.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
    dash_table_row(ws_d, i, [
        (2, interes, None),
        (5, cnt, "#,##0"),
        (6, ins, "#,##0"),
        (7, ins/cnt if cnt else 0, "0.0%"),
        (8, cnt/total if total else 0, "0.0%"),
    ], alt=i%2==0)

# ── ESTATUS DISTRIBUCIÓN TABLE ─────────────────────────────────────────────
dash_section_label(ws_d, START_INT, 10, 16, "📊  DISTRIBUCIÓN POR ESTATUS", C_ACC2)
dash_table_header(ws_d, START_INT+1,
    [(10,"ESTATUS FINAL"),(13,"TOTAL"),(14,"% LEADS"),(15,"OBSERVACIÓN")], C_DARK)
ws_d.merge_cells(f"J{START_INT+1}:L{START_INT+1}")

ESTATUS_OBS = {
    "6 - Inscrito":              "✅ Convertido",
    "2 - Interesado":            "🔥 Trabajar ahora",
    "5 - En proceso inscripción":"⏳ Cerrar pronto",
    "3 - Contactado/Activo":     "👀 En seguimiento",
    "4 - Clase muestra":         "🏃 Mostrar valor",
    "7 - Perdido":               "❌ No convertido",
    "8 - Sin respuesta":         "📵 Reactivar",
    "1 - Lead/Informes":         "📥 Por contactar",
}
estatus_sorted = sorted(by_estatus.items(), key=lambda x: -x[1])
for i, (est, cnt) in enumerate(estatus_sorted, START_INT+2):
    obs = ESTATUS_OBS.get(est, "")
    est_bg = {
        "6 - Inscrito": "D4EDDA",
        "7 - Perdido": "F8D7DA",
        "8 - Sin respuesta": "FFF3CD",
        "2 - Interesado": "CCE5FF",
    }.get(est, C_LGRAY)
    ws_d.merge_cells(start_row=i, start_column=10, end_row=i, end_column=12)
    for col in [10, 13, 14, 15]:
        ws_d.cell(row=i, column=col).fill = fill(est_bg)
        ws_d.cell(row=i, column=col).border = border_all()
        ws_d.cell(row=i, column=col).alignment = align()
        ws_d.cell(row=i, column=col).font = Font(name="Calibri", size=9, color=C_TEXT)
    ws_d.cell(row=i, column=10).value = est
    ws_d.cell(row=i, column=13).value = cnt
    ws_d.cell(row=i, column=13).number_format = "#,##0"
    ws_d.cell(row=i, column=14).value = cnt/total if total else 0
    ws_d.cell(row=i, column=14).number_format = "0.0%"
    ws_d.cell(row=i, column=15).value = obs

print("Dashboard sections ok")


# ── Footer ─────────────────────────────────────────────────────────────────
last_row = START_INT + len(estatus_sorted) + 4
ws_d.merge_cells(start_row=last_row, start_column=2, end_row=last_row, end_column=20)
footer = ws_d.cell(row=last_row, column=2,
    value="⚡ Para actualizar el dashboard, agrega leads en LEADS_DB  ·  Columnas N, O, M se calcularán automáticamente con Apps Script")
footer.font = Font(name="Calibri", size=8, color=C_GRAY, italic=True)
footer.fill = fill(C_MID)
footer.alignment = align()

# ── Chart 1: Funnel bar chart ─────────────────────────────────────────────
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList

chart1 = BarChart()
chart1.type = "bar"
chart1.title = "Embudo de Conversión"
chart1.style = 10
chart1.y_axis.title = "Etapa"
chart1.x_axis.title = "Leads"
chart1.width = 16
chart1.height = 10
chart1.grouping = "clustered"

# Build a small helper sheet for chart data
ws_chart = wb.create_sheet("_ChartData")
ws_chart.sheet_state = "hidden"
funnel_labels = [f[0] for f in funnel]
funnel_vals   = [f[1] for f in funnel]
ws_chart["A1"] = "Etapa"; ws_chart["B1"] = "Leads"; ws_chart["C1"] = "Inscritos"
for i, (lbl, val) in enumerate(zip(funnel_labels, funnel_vals), 2):
    ws_chart.cell(row=i, column=1, value=lbl)
    ws_chart.cell(row=i, column=2, value=val)
ws_chart["C2"] = inscritos

# Fuente chart data
ws_chart["E1"] = "Fuente"; ws_chart["F1"] = "Leads"; ws_chart["G1"] = "Inscritos"
for i, (f, c) in enumerate(fuentes_sorted, 2):
    ws_chart.cell(row=i, column=5, value=f)
    ws_chart.cell(row=i, column=6, value=c)
    ws_chart.cell(row=i, column=7, value=ins_by_fuente.get(f, 0))

# Monthly chart data
ws_chart["I1"] = "Mes"; ws_chart["J1"] = "Leads"; ws_chart["K1"] = "Inscritos"
for i, mes in enumerate(meses_sorted, 2):
    ws_chart.cell(row=i, column=9, value=mes)
    ws_chart.cell(row=i, column=10, value=by_mes[mes])
    ws_chart.cell(row=i, column=11, value=ins_by_mes.get(mes, 0))

# Funnel bar chart
n_funnel = len(funnel) + 1
data_f = Reference(ws_chart, min_col=2, max_col=2, min_row=1, max_row=n_funnel)
cats_f = Reference(ws_chart, min_col=1, min_row=2, max_row=n_funnel)
chart1.add_data(data_f, titles_from_data=True)
chart1.set_categories(cats_f)
chart1.series[0].graphicalProperties.solidFill = "0F3460"
ws_d.add_chart(chart1, "B26")

# Fuente bar chart
n_fuente = len(fuentes_sorted) + 1
chart2 = BarChart()
chart2.type = "bar"
chart2.title = "Leads por Fuente"
chart2.style = 10
chart2.width = 16; chart2.height = 10
data_fu = Reference(ws_chart, min_col=6, max_col=7, min_row=1, max_row=n_fuente)
cats_fu = Reference(ws_chart, min_col=5, min_row=2, max_row=n_fuente)
chart2.add_data(data_fu, titles_from_data=True)
chart2.set_categories(cats_fu)
chart2.series[0].graphicalProperties.solidFill = "E94560"
chart2.series[1].graphicalProperties.solidFill = "00B894"
ws_d.add_chart(chart2, "J26")

# Monthly line chart
from openpyxl.chart import LineChart
n_mes = len(meses_sorted) + 1
chart3 = LineChart()
chart3.title = "Evolución Mensual"
chart3.style = 10
chart3.width = 22; chart3.height = 12
data_m = Reference(ws_chart, min_col=10, max_col=11, min_row=1, max_row=n_mes)
cats_m = Reference(ws_chart, min_col=9, min_row=2, max_row=n_mes)
chart3.add_data(data_m, titles_from_data=True)
chart3.set_categories(cats_m)
chart3.series[0].graphicalProperties.line.solidFill = "0F3460"
chart3.series[0].graphicalProperties.line.width = 25000
chart3.series[1].graphicalProperties.line.solidFill = "00B894"
chart3.series[1].graphicalProperties.line.width = 25000
ws_d.add_chart(chart3, "B42")

print("Charts added")

# ══════════════════════════════════════════════════════════════════════════════
# Save
# ══════════════════════════════════════════════════════════════════════════════
wb.save(DST)
print(f"\n✅ Archivo guardado: {DST}")
print(f"   Hojas: {[s.title for s in wb.worksheets]}")
print(f"   Leads procesados: {total}")
print(f"   Inscritos: {inscritos}  ({tasa_conv:.1%} conversión)")
