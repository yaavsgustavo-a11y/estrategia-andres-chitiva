"""
build_base_leads.py
Genera BASE_LEADS_CHITIVA_LIMPIO.xlsx desde "BASE LEADS CHITIVA JULIO .xlsx"
  - Limpieza: años 2028→2026, elimina filas sin ningún dato útil
  - Homologación: Interés, Medio, Fuente, Canal, Etapa, Resultado
  - Valores múltiples (ej. "Academia / Curso de Verano") → soportados con separador " / "
  - Dropdowns de validación de datos en todas las columnas clave
  - ID autoincremental
  - Columnas auxiliares: Semana, Mes
  - Hoja Config con listas
  - Hoja DATOS_PIVOT con COUNTIF/COUNTIFS como fórmulas (se actualizan solas)
  - Hoja DASHBOARD con KPIs y gráficos vinculados a DATOS_PIVOT
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, date
from collections import Counter, defaultdict
import re

SRC = "BASE LEADS CHITIVA JULIO .xlsx"
DST = "BASE_LEADS_CHITIVA_LIMPIO.xlsx"

# ── Paleta ─────────────────────────────────────────────────────────────────
C_DARK   = "1A1A2E"; C_MID = "16213E"; C_ACC1 = "0F3460"
C_ACC2   = "E94560"; C_GREEN = "00B894"; C_YELLOW = "FDCB6E"
C_ORANGE = "E17055"; C_WHITE = "FFFFFF"; C_LGRAY = "F0F2F5"
C_GRAY   = "BDC3C7"; C_TEXT  = "2D3436"; C_BLUE  = "74B9FF"

def fill(h): return PatternFill("solid", fgColor=h)
def fnt(bold=False, sz=10, color=C_TEXT, italic=False):
    return Font(name="Calibri", bold=bold, size=sz, color=color, italic=italic)
def aln(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr(c=C_GRAY):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def thick_bdr(c=C_ACC1):
    s = Side(style="medium", color=c)
    return Border(left=s, right=s, top=s, bottom=s)

# ══════════════════════════════════════════════════════════════════════════
# TABLAS DE HOMOLOGACIÓN
# ══════════════════════════════════════════════════════════════════════════

# Interés — valores canónicos y sus alias
# IMPORTANTE: un registro puede tener 2 intereses → los separamos con " / "
INTERES_CANONICO = {
    # Academia de Fútbol
    "ACADEMIA": "Academia de Fútbol",
    "ACADEMIA DE FUTBOL": "Academia de Fútbol",
    "ACADEMIA DE FUTBOLL": "Academia de Fútbol",
    "ACADEMIA DE FUTBOLL Y LIGAS": "Academia de Fútbol",
    "ACADE": "Academia de Fútbol",
    "ACADEMIA DE FUTBOL Y LIGAS": "Academia de Fútbol",
    # Curso de Verano
    "CURSO DE VERANO": "Curso de Verano",
    "CURSO": "Curso de Verano",
    "CU": "Curso de Verano",
    "CURSO PILOTO": "Curso de Verano",
    # Torneo
    "TORNEO": "Torneo",
    "TOR": "Torneo",
    "TORNEO INFANTIL": "Torneo Infantil",
    "TORNEO FIN DE SEMANA": "Torneo Fin de Semana",
    # Ligas
    "SEGUNDA VARONIL": "Liga 2a Varonil",
    "TERCERA VARONIL": "Liga 3a Varonil",
    "PREMIER VARONIL": "Liga Premier Varonil",
    "PRIMERA, SEGUNDA Y TERCERA": "Ligas Varoniles",
    "SEGUNDA Y TERCERA VARONIL": "Ligas Varoniles",
    "SEGUNDA O TERCERA VARONIL": "Ligas Varoniles",
    "SEGUNDA Y PREMIER VARONIL": "Ligas Varoniles",
    "PREMIER Y TERCERA VARONIL": "Ligas Varoniles",
    "TERCERA Y FIN DE SEMANA VARONIL": "Liga 3a Varonil",
    "SEGUNDA Y FIN DE SEMANA VARONIL": "Liga 2a Varonil",
    "LIGA VARONIL": "Ligas Varoniles",
    "PREMIER FEMENIL": "Liga Premier Femenil",
    "SEGUNDA FEMENIL": "Liga 2a Femenil",
    "LIGAS INFANTILES": "Liga Infantil",
    "LIGA INFANTIL": "Liga Infantil",
    "LIGA VETERANOS": "Liga Veteranos",
    "INFANTILES": "Liga Infantil",
    "INFANTIL": "Liga Infantil",
    "VETERANO": "Liga Veteranos",
    "VETERANOS": "Liga Veteranos",
    "VETERANOS VARONIL": "Liga Veteranos",
    # Pretemporada / Entrenamiento
    "PRETEMPORADA INTENSIVA": "Pretemporada Intensiva",
    "ENTRENAMIENTO FUNCIONAL": "Entrenamiento Funcional",
    "ENTRENAMIENTO DE PORTEROS": "Entrenamiento Funcional",
    # Fin de Semana
    "FIN DE SEMANA": "Fin de Semana",
    "FIN DE SEMANA VARONIL": "Fin de Semana",
    # Pádel
    "PADEL": "Pádel",
    # Empresarial
    "EMPRESARIAL": "Empresarial",
    # General
    "GENERAL": "General",
    "UBICACIÓN": "General",
    # Renta
    "RENTA DE CANCHA": "Renta de Cancha",
    "RENTA DE CANCHAS": "Renta de Cancha",
    "RENTA CANCHA": "Renta de Cancha",
    # Otros
    "PARTIDO AMISTOSO": "Partido Amistoso",
    "REHABILITACIÓN": "Rehabilitación",
}

# Combinaciones multi-interés (texto original → lista de valores canónicos)
INTERES_MULTI = {
    "CURSO DE VERANO / ENTRENAMIENTO": ["Curso de Verano", "Entrenamiento Funcional"],
    "CURSO DE VERANO / PADEL":         ["Curso de Verano", "Pádel"],
    "GENERAL / CURSO DE VERANO":       ["General", "Curso de Verano"],
    "ACADEMIA Y CURSO":                ["Academia de Fútbol", "Curso de Verano"],
    "ACADEMIA Y CU":                   ["Academia de Fútbol", "Curso de Verano"],
    "ACADEMIA Y CU":                   ["Academia de Fútbol", "Curso de Verano"],
    "CURSO Y ACADEMIA":                ["Curso de Verano", "Academia de Fútbol"],
    "ACADEMIA Y CURSO":                ["Academia de Fútbol", "Curso de Verano"],
    "ACADEMIA, CURSOS":                ["Academia de Fútbol", "Curso de Verano"],
    "ACADEMIA Y CURSOS":               ["Academia de Fútbol", "Curso de Verano"],
    "FIN DE SEMANA VARONIL / ACADEMIA":["Fin de Semana", "Academia de Fútbol"],
    "ACADEMIA y curso":                ["Academia de Fútbol", "Curso de Verano"],
    "academia y cu":                   ["Academia de Fútbol", "Curso de Verano"],
    "ACADEMIA Y CURSO":                ["Academia de Fútbol", "Curso de Verano"],
}

def normalizar_interes(raw):
    if not raw or str(raw).strip() in ("", "-", "None"):
        return ""
    v = str(raw).strip()
    # Revisar multi primero
    key = v.strip()
    for orig, vals in INTERES_MULTI.items():
        if key.upper() == orig.upper():
            return " / ".join(vals)
    # Intentar lookup directo
    up = key.upper()
    if up in INTERES_CANONICO:
        return INTERES_CANONICO[up]
    # Búsqueda parcial
    if "ACADEMIA" in up:
        if "CURSO" in up:
            return "Academia de Fútbol / Curso de Verano"
        return "Academia de Fútbol"
    if "CURSO" in up:
        if "ACADEMIA" in up:
            return "Academia de Fútbol / Curso de Verano"
        if "PADEL" in up or "PÁDEL" in up:
            return "Curso de Verano / Pádel"
        if "ENTRENAMIENTO" in up:
            return "Curso de Verano / Entrenamiento Funcional"
        return "Curso de Verano"
    if "TORNEO" in up:
        if "INFANTIL" in up:
            return "Torneo Infantil"
        if "FIN DE SEMANA" in up:
            return "Torneo Fin de Semana"
        return "Torneo"
    if "TOR" == up:
        return "Torneo"
    if "LIGA" in up or "VARONIL" in up or "FEMENIL" in up:
        if "INFANTIL" in up:
            return "Liga Infantil"
        if "VETERANO" in up:
            return "Liga Veteranos"
        if "PREMIER" in up:
            return "Liga Premier Varonil"
        if "SEGUNDA" in up and "TERCERA" in up:
            return "Ligas Varoniles"
        if "SEGUNDA" in up:
            return "Liga 2a Varonil"
        if "TERCERA" in up:
            return "Liga 3a Varonil"
        return "Ligas Varoniles"
    if "ENTRENAMIENTO" in up or "FUNCIONAL" in up or "PORTERO" in up:
        return "Entrenamiento Funcional"
    if "PRETEMPORADA" in up:
        return "Pretemporada Intensiva"
    if "FIN DE SEMANA" in up:
        return "Fin de Semana"
    if "PADEL" in up or "PÁDEL" in up:
        return "Pádel"
    if "EMPRESARIAL" in up:
        return "Empresarial"
    if "VETERANO" in up or "VETERANOS" in up:
        return "Liga Veteranos"
    if "RENTA" in up:
        return "Renta de Cancha"
    if "GENERAL" in up or "UBICACI" in up:
        return "General"
    if "REHABILIT" in up:
        return "Rehabilitación"
    if "PARTIDO" in up or "AMISTOSO" in up:
        return "Partido Amistoso"
    if "INFANTIL" in up:
        return "Liga Infantil"
    return v  # devolver sin cambios si no se reconoce

MEDIO_MAP = {
    "FACEBOOK": "Facebook", "facebook": "Facebook",
    "FACEBOK": "Facebook", "facebok": "Facebook",
    "FACEBO": "Facebook",
    "WHATSAPP": "WhatsApp", "whatsapp": "WhatsApp",
    "INSTAGRAM": "Instagram",
    "TIK TOK": "TikTok", "tik tok": "TikTok",
    "TIKTOK": "TikTok", "tikTOK": "TikTok",
    "MESSENGER": "Messenger",
    "LLAMADA": "Llamada",
    "OFICINAS": "Oficinas",
    "CERCANIA": "Cercanía",
    "CHITIVA": "Referido/Conocido",
}
def norm_medio(v):
    if not v or str(v).strip() in ("","-","None"): return ""
    return MEDIO_MAP.get(str(v).strip(), MEDIO_MAP.get(str(v).strip().upper(), str(v).strip().title()))

FUENTE_MAP = {
    "FACEBOOK": "Facebook", "facebook": "Facebook",
    "WHATSAPP": "WhatsApp", "whatsapp": "WhatsApp",
    "whatSAPP": "WhatsApp", "whatsAPP": "WhatsApp",
    "WW": "WhatsApp", "WHS": "WhatsApp",
    "CATELOGO DE WHATSAPP": "WhatsApp/Catálogo",
    "CATALOGO DE WHATSAPP": "WhatsApp/Catálogo",
    "INSTAGRAM": "Instagram",
    "TIK TOK": "TikTok", "tik tok": "TikTok",
    "TIKTOK": "TikTok", "tikTOK": "TikTok",
    "CONOCIDO": "Referido/Conocido",
    "CERCANIA DE INSTALACIONES": "Cercanía",
    "CERCANIA": "Cercanía",
    "VIDEO": "Orgánico",
    "ACADEMIA": "Cercanía",
    "OFICINAS": "Oficinas",
    "FACEBO": "Facebook",
    "FACEBOK": "Facebook",
}
def norm_fuente(v):
    if not v or str(v).strip() in ("","-","None"): return ""
    s = str(v).strip()
    if s in FUENTE_MAP: return FUENTE_MAP[s]
    up = s.upper()
    if "FACEBOOK" in up: return "Facebook"
    if "INSTAGRAM" in up: return "Instagram"
    if "TIKTOK" in up or "TIK TOK" in up: return "TikTok"
    if "WHATSAPP" in up or "WHAS" in up or "CATALO" in up: return "WhatsApp"
    if "CONOCIDO" in up or "REFERIDO" in up: return "Referido/Conocido"
    if "CERCANIA" in up or "CERCANÍA" in up: return "Cercanía"
    return s.title()

def norm_canal(v):
    if not v or str(v).strip() in ("","-","None"): return ""
    up = str(v).strip().upper()
    if "WHATSAPP" in up or "WHAS" in up or "WHS" in up: return "WhatsApp"
    if "OFICINAS" in up: return "Oficinas"
    if "LLAMADA" in up: return "Llamada"
    return "WhatsApp"

ETAPA_MAP_RULES = [
    (["COMPLETO","INSCRITO"], "Completo"),
    (["PROCESO DE INSCRIPCI"], "En Proceso de Inscripción"),
    (["INTERESADO"], "En Informes / Interesado"),
    (["INFORMES"], "En Informes / Interesado"),
    (["SIN CONTESTACI","NO CONTESTO"], "Sin Contestación"),
    (["SIN RESPUESTA"], "Sin Contestación"),
    (["ESPERA DE RECOLECCIÓN","ESPERA DE RECOLECCION","REGISTRO DE INFORM","ESPERA"], "En Espera"),
    (["PARTIDO AMISTOSO","TRATO"], "Contactado/Activo"),
    (["PROCESO DE INSCRIPCION"], "En Proceso de Inscripción"),
]
def norm_etapa(v):
    if not v or str(v).strip() in ("","-","None"): return ""
    up = str(v).strip().upper()
    for keywords, canonical in ETAPA_MAP_RULES:
        if any(k in up for k in keywords):
            return canonical
    return str(v).strip().title()

RESULTADO_MAP_RULES = [
    (["INSCRITO"], "Inscrito"),
    (["PROCESO DE PAGO","PROCESO DE INSCRIPCI"], "En Proceso de Inscripción"),
    (["CLASE MUESTRA","RESPUESTA DEPUES DE CLASE MUESTRA"], "Clase Muestra"),
    (["PARTIDO AMISTOSO","PROXIMA VISITA","SIGUIENTE SEMANA","TRATO","CONVENIO","RENTA DE CANCHA","ENTRENAMIENTO DE PORTEROS"], "Contactado/Activo"),
    (["ELECCIÓN CURSO","ELECCION CURSO"], "Interesado"),
    (["INTERESADA","INTERESADO"], "Interesado"),
    (["ESPERANDO RESPUESTA","ESPERA DE RESPUESTA","EN ESPERA","ESPERA DE LIGA","ESPERA DE PAGO","ESPERA DE INSCRIPCI","ESPERA DE PROCESO","ESPERA"], "En Espera"),
    (["DESINTERESADO","NO INSCRITO","SIN LIGAS","NO HAY AUN TORNEO","NO ESTA SU HIJO","NO TENEMOS LIGA","NO TIENE EQUIPO"], "Desinteresado"),
    (["DEJO DE CONTESTAR","NO RESPONDIO","NO VOLVIO A CONTESTAR","SIN RESPUESTA","SIN CONTESTACI"], "Sin Respuesta"),
    (["COMPLETO"], "Inscrito"),
    (["INFORMES"], "Interesado"),
]
def norm_resultado(v):
    if not v or str(v).strip() in ("","-","None","0.9"): return ""
    up = str(v).strip().upper()
    for keywords, canonical in RESULTADO_MAP_RULES:
        if any(k in up for k in keywords):
            return canonical
    return str(v).strip().title()

def get_semana(d):
    if isinstance(d, datetime): d = d.date()
    if not isinstance(d, date): return ""
    iso = d.isocalendar()
    return f"{iso[0]}-S{iso[1]:02d}"

def get_mes(d):
    if isinstance(d, datetime): d = d.date()
    if not isinstance(d, date): return ""
    return d.strftime("%Y-%m")

def fix_fecha(v):
    if isinstance(v, datetime):
        if v.year == 2028: return v.replace(year=2026)
        if v.year == 2027: return v.replace(year=2026)
        return v
    return v

def nonempty(v):
    return v is not None and str(v).strip() not in ("", "-", "None")


# ══════════════════════════════════════════════════════════════════════════
# PASO 1 — LEER Y LIMPIAR DATOS FUENTE
# ══════════════════════════════════════════════════════════════════════════
print("▶ Leyendo fuente…")
wb_src = openpyxl.load_workbook(SRC, data_only=True)
ws_src = wb_src["Hoja1"]

leads = []   # filas con nombre O teléfono
for row in range(3, ws_src.max_row + 1):
    vals = [ws_src.cell(row, c).value for c in range(1, 12)]
    fecha_raw, nombre, telefono, correo, interes, medio, anuncio, canal, etapa, resultado, comentario = vals

    # Decidir si la fila tiene algo útil
    tiene_nombre = nonempty(nombre)
    tiene_tel    = nonempty(telefono)
    tiene_fecha  = fecha_raw is not None
    tiene_alguno = any(nonempty(v) for v in vals)

    # Eliminar si no hay absolutamente nada
    if not tiene_alguno:
        continue
    # Eliminar si solo hay fecha sin ningún otro dato
    if tiene_fecha and not tiene_nombre and not tiene_tel and not any(nonempty(v) for v in vals[3:]):
        continue

    fecha = fix_fecha(fecha_raw)

    leads.append({
        "fecha":     fecha.date() if isinstance(fecha, datetime) else fecha,
        "nombre":    str(nombre).strip() if tiene_nombre else "",
        "telefono":  str(telefono).replace(".0","").strip() if tiene_tel else "",
        "correo":    str(correo).strip() if nonempty(correo) else "",
        "interes_raw": str(interes).strip() if nonempty(interes) else "",
        "medio_raw":   str(medio).strip()   if nonempty(medio)   else "",
        "anuncio_raw": str(anuncio).strip() if nonempty(anuncio) else "",
        "canal_raw":   str(canal).strip()   if nonempty(canal)   else "",
        "etapa_raw":   str(etapa).strip()   if nonempty(etapa)   else "",
        "resultado_raw": str(resultado).strip() if nonempty(resultado) else "",
        "comentario":  str(comentario).strip() if nonempty(comentario) else "",
    })

# ── Aplicar homologaciones ────────────────────────────────────────────────
for i, l in enumerate(leads):
    l["id"]        = i + 1
    l["interes"]   = normalizar_interes(l["interes_raw"])
    l["medio"]     = norm_medio(l["medio_raw"])
    l["fuente"]    = norm_fuente(l["anuncio_raw"]) or norm_fuente(l["medio_raw"])
    l["canal"]     = norm_canal(l["canal_raw"])
    l["etapa"]     = norm_etapa(l["etapa_raw"])
    l["resultado"] = norm_resultado(l["resultado_raw"])
    l["semana"]    = get_semana(l["fecha"])
    l["mes"]       = get_mes(l["fecha"])

total = len(leads)
con_nombre = sum(1 for l in leads if l["nombre"])
sin_nombre = total - con_nombre
print(f"   Total filas útiles: {total}  (con nombre: {con_nombre}, solo teléfono: {sin_nombre})")

# ── Listas canónicas para Config / dropdowns ──────────────────────────────
LISTA_INTERESES = sorted(set([
    "Academia de Fútbol", "Curso de Verano", "Torneo", "Torneo Infantil",
    "Torneo Fin de Semana", "Liga 2a Varonil", "Liga 3a Varonil",
    "Liga Premier Varonil", "Liga Premier Femenil", "Liga 2a Femenil",
    "Liga Infantil", "Liga Veteranos", "Ligas Varoniles",
    "Pretemporada Intensiva", "Entrenamiento Funcional", "Fin de Semana",
    "Pádel", "Empresarial", "General", "Renta de Cancha",
    "Partido Amistoso", "Rehabilitación",
]))
LISTA_MULTI_INTERESES = [
    "Academia de Fútbol / Curso de Verano",
    "Curso de Verano / Pádel",
    "Curso de Verano / Entrenamiento Funcional",
    "Fin de Semana / Academia de Fútbol",
    "General / Curso de Verano",
]

LISTA_MEDIO   = ["Facebook","WhatsApp","Instagram","TikTok","Messenger","Llamada","Oficinas","Cercanía","Referido/Conocido"]
LISTA_FUENTE  = ["Facebook","Instagram","TikTok","WhatsApp","WhatsApp/Catálogo","Referido/Conocido","Cercanía","Orgánico","Oficinas","Otro"]
LISTA_CANAL   = ["WhatsApp","Llamada","Oficinas","Instagram","Facebook","Messenger"]
LISTA_ETAPA   = ["En Informes / Interesado","En Proceso de Inscripción","Completo","Sin Contestación","En Espera","Contactado/Activo"]
LISTA_RESULTADO = [
    "Inscrito","En Proceso de Inscripción","Clase Muestra","Contactado/Activo",
    "Interesado","En Espera","Desinteresado","Sin Respuesta",
]

# ── Colores por resultado (para la columna resultado en LEADS_DB) ─────────
RESULTADO_COLOR = {
    "Inscrito":                  (C_GREEN,  C_WHITE),
    "En Proceso de Inscripción": (C_BLUE,   C_WHITE),
    "Clase Muestra":             ("A29BFE", C_WHITE),
    "Contactado/Activo":         (C_ACC1,   C_WHITE),
    "Interesado":                (C_YELLOW, C_TEXT),
    "En Espera":                 (C_ORANGE, C_WHITE),
    "Desinteresado":             (C_ACC2,   C_WHITE),
    "Sin Respuesta":             (C_GRAY,   C_TEXT),
}


# ══════════════════════════════════════════════════════════════════════════
# CREAR WORKBOOK DE SALIDA
# ══════════════════════════════════════════════════════════════════════════
print("▶ Creando workbook de salida…")
wb = Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════════════════
# HOJA: Config  (listas para validación de datos)
# ══════════════════════════════════════════════════════════════════════════
ws_cfg = wb.create_sheet("Config")
ws_cfg.sheet_properties.tabColor = "636E72"
ws_cfg.sheet_state = "visible"   # visible para que el usuario la vea

def cfg_block(ws, col, title, items, bg=C_ACC1):
    ws.column_dimensions[get_column_letter(col)].width = 32
    hdr = ws.cell(1, col, title)
    hdr.font = fnt(bold=True, sz=10, color=C_WHITE)
    hdr.fill = fill(bg)
    hdr.alignment = aln()
    hdr.border = bdr(bg)
    for i, item in enumerate(items, 2):
        c = ws.cell(i, col, item)
        c.font = fnt(sz=9)
        c.fill = fill(C_LGRAY if i % 2 == 0 else C_WHITE)
        c.border = bdr()
        c.alignment = aln(h="left")

cfg_block(ws_cfg, 1, "INTERÉS (simple)",   LISTA_INTERESES,         C_ACC1)
cfg_block(ws_cfg, 2, "INTERÉS (múltiple)", LISTA_MULTI_INTERESES,   C_MID)
cfg_block(ws_cfg, 3, "MEDIO",              LISTA_MEDIO,             C_ACC2)
cfg_block(ws_cfg, 4, "FUENTE/ANUNCIO",     LISTA_FUENTE,            C_MID)
cfg_block(ws_cfg, 5, "CANAL SEGUIMIENTO",  LISTA_CANAL,             C_ACC1)
cfg_block(ws_cfg, 6, "ETAPA",              LISTA_ETAPA,             C_ACC2)
cfg_block(ws_cfg, 7, "RESULTADO",          LISTA_RESULTADO,         C_GREEN)

# Combinar listas interés para dropdown
LISTA_INTERES_COMPLETA = LISTA_INTERESES + LISTA_MULTI_INTERESES
cfg_block(ws_cfg, 9, "INTERÉS (completa)", LISTA_INTERES_COMPLETA,  C_MID)

print("   ✓ Config")

# ══════════════════════════════════════════════════════════════════════════
# HOJA: LEADS_DB
# ══════════════════════════════════════════════════════════════════════════
ws_db = wb.create_sheet("LEADS_DB")
ws_db.sheet_properties.tabColor = "0F3460"

# Columnas
DB_COLS = [
    ("ID",                5),
    ("FECHA",            12),
    ("NOMBRE",           28),
    ("TELÉFONO",         16),
    ("CORREO",           24),
    ("INTERÉS",          30),
    ("MEDIO",            16),
    ("FUENTE/ANUNCIO",   20),
    ("CANAL",            14),
    ("ETAPA",            22),
    ("RESULTADO",        22),
    ("COMENTARIO",       30),
    ("SEMANA",           10),
    ("MES",              10),
]
NCOLS = len(DB_COLS)

for ci, (_, w) in enumerate(DB_COLS, 1):
    ws_db.column_dimensions[get_column_letter(ci)].width = w

ws_db.freeze_panes = "A3"
ws_db.row_dimensions[1].height = 32
ws_db.row_dimensions[2].height = 22

# Fila 1 — título
ws_db.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
tc = ws_db["A1"]
tc.value = "📋  BASE DE DATOS LEADS — ANDRES CHITIVA"
tc.font  = fnt(bold=True, sz=14, color=C_WHITE)
tc.fill  = fill(C_DARK)
tc.alignment = aln()

# Fila 2 — encabezados
for ci, (hdr, _) in enumerate(DB_COLS, 1):
    c = ws_db.cell(2, ci, hdr)
    c.font      = fnt(bold=True, sz=9, color=C_WHITE)
    c.fill      = fill(C_ACC1)
    c.alignment = aln(wrap=True)
    c.border    = bdr(C_ACC1)

# ── Filas de datos ────────────────────────────────────────────────────────
ALT = [fill(C_LGRAY), fill(C_WHITE)]
for i, lead in enumerate(leads, 1):
    r = i + 2
    ws_db.row_dimensions[r].height = 15
    bg = ALT[i % 2]

    row_vals = [
        lead["id"],
        lead["fecha"],
        lead["nombre"],
        lead["telefono"],
        lead["correo"],
        lead["interes"],
        lead["medio"],
        lead["fuente"],
        lead["canal"],
        lead["etapa"],
        lead["resultado"],
        lead["comentario"],
        lead["semana"],
        lead["mes"],
    ]

    for ci, v in enumerate(row_vals, 1):
        cell = ws_db.cell(r, ci, v)
        cell.font      = fnt(sz=9)
        cell.fill      = bg
        cell.border    = bdr()
        cell.alignment = aln(h="center" if ci in (1,2,7,8,9,10,11,13,14) else "left")
        if ci == 2 and isinstance(v, date):
            cell.number_format = "DD/MM/YYYY"

    # Color en columna RESULTADO (col 11)
    res_cell = ws_db.cell(r, 11)
    res_val  = lead["resultado"]
    if res_val in RESULTADO_COLOR:
        bg_c, fg_c = RESULTADO_COLOR[res_val]
        res_cell.fill = fill(bg_c)
        res_cell.font = fnt(sz=9, bold=True, color=fg_c)

    # Color en columna ETAPA (col 10) — diferente tono
    etapa_cell = ws_db.cell(r, 10)
    etapa_val  = lead["etapa"]
    ETAPA_COLOR = {
        "Completo":                   (C_GREEN,  C_WHITE),
        "En Proceso de Inscripción":  (C_BLUE,   C_WHITE),
        "En Informes / Interesado":   ("DFE6E9",  C_TEXT),
        "Sin Contestación":           (C_ORANGE, C_WHITE),
        "En Espera":                  (C_YELLOW, C_TEXT),
        "Contactado/Activo":          (C_MID,    C_WHITE),
    }
    if etapa_val in ETAPA_COLOR:
        eb, ef = ETAPA_COLOR[etapa_val]
        etapa_cell.fill = fill(eb)
        etapa_cell.font = fnt(sz=9, color=ef)

# ── Validaciones de datos (dropdowns) ────────────────────────────────────
last_data_row = len(leads) + 200   # margen para nuevas entradas

def make_dv(formula, prompt_title, prompt_body):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    dv.prompt        = prompt_body
    dv.promptTitle   = prompt_title
    dv.showInputMessage = True
    return dv

# Rangos en Config — usamos columna 9 (interés completa)
n_int  = len(LISTA_INTERES_COMPLETA) + 1
n_med  = len(LISTA_MEDIO) + 1
n_fue  = len(LISTA_FUENTE) + 1
n_can  = len(LISTA_CANAL) + 1
n_eta  = len(LISTA_ETAPA) + 1
n_res  = len(LISTA_RESULTADO) + 1

dv_interes   = make_dv(f"Config!$I$2:$I${n_int}",  "Interés",    "Selecciona el tipo de interés")
dv_medio     = make_dv(f"Config!$C$2:$C${n_med}",  "Medio",      "Canal por donde contactó")
dv_fuente    = make_dv(f"Config!$D$2:$D${n_fue}",  "Fuente",     "Origen/anuncio del lead")
dv_canal     = make_dv(f"Config!$E$2:$E${n_can}",  "Canal",      "Canal de seguimiento")
dv_etapa     = make_dv(f"Config!$F$2:$F${n_eta}",  "Etapa",      "Etapa actual del proceso")
dv_resultado = make_dv(f"Config!$G$2:$G${n_res}",  "Resultado",  "Resultado del seguimiento")

ws_db.add_data_validation(dv_interes)
ws_db.add_data_validation(dv_medio)
ws_db.add_data_validation(dv_fuente)
ws_db.add_data_validation(dv_canal)
ws_db.add_data_validation(dv_etapa)
ws_db.add_data_validation(dv_resultado)

dv_interes.sqref   = f"F3:F{last_data_row}"
dv_medio.sqref     = f"G3:G{last_data_row}"
dv_fuente.sqref    = f"H3:H{last_data_row}"
dv_canal.sqref     = f"I3:I{last_data_row}"
dv_etapa.sqref     = f"J3:J{last_data_row}"
dv_resultado.sqref = f"K3:K{last_data_row}"

print(f"   ✓ LEADS_DB  ({len(leads)} filas, {NCOLS} columnas, dropdowns OK)")


# ══════════════════════════════════════════════════════════════════════════
# HOJA: DATOS_PIVOT  — todo fórmulas COUNTIF/COUNTIFS, se actualiza solo
# ══════════════════════════════════════════════════════════════════════════
print("▶ Creando DATOS_PIVOT…")
ws_pv = wb.create_sheet("DATOS_PIVOT")
ws_pv.sheet_properties.tabColor = "0984E3"

DB = "LEADS_DB"   # nombre de la hoja fuente

def pv_title(ws, row, col, text, span=4, bg=C_DARK):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    c = ws.cell(row, col, text)
    c.font = fnt(bold=True, sz=11, color=C_WHITE)
    c.fill = fill(bg)
    c.alignment = aln()

def pv_hdr(ws, row, col, text, bg=C_ACC1):
    c = ws.cell(row, col, text)
    c.font = fnt(bold=True, sz=9, color=C_WHITE)
    c.fill = fill(bg)
    c.alignment = aln()
    c.border = bdr(bg)

def pv_cell(ws, row, col, val_or_formula, fmt=None, bg=C_WHITE, bold=False):
    c = ws.cell(row, col, val_or_formula)
    c.font = fnt(sz=9, bold=bold)
    c.fill = fill(bg)
    c.border = bdr()
    c.alignment = aln()
    if fmt: c.number_format = fmt
    return c

def pv_label(ws, row, col, text, bg=C_LGRAY):
    c = ws.cell(row, col, text)
    c.font = fnt(sz=9)
    c.fill = fill(bg)
    c.border = bdr()
    c.alignment = aln(h="left")

# Anchos de columna
for col, w in [(1,28),(2,12),(3,14),(4,14),(5,2),(6,28),(7,12),(8,14),(9,14),(10,2),(11,14),(12,14),(13,14),(14,14)]:
    ws_pv.column_dimensions[get_column_letter(col)].width = w


# ── Sección A: KPIs Generales (col 1) ─────────────────────────────────────
pv_title(ws_pv, 1, 1, "📊  KPIs GENERALES", 4, C_DARK)
for ci, h in [(1,"MÉTRICA"),(2,"VALOR"),(3,"DETALLE"),(4,"")]:
    pv_hdr(ws_pv, 2, ci, h, C_MID)

# Conteo total = filas con nombre O teléfono (col C o D, fila 3 en adelante)
KPI_ROWS = [
    ("Total Leads",           f'=COUNTA({DB}!C3:C5000)+COUNTBLANK({DB}!C3:C5000)-COUNTIF({DB}!C3:C5000,"")',  ""),
    ("Con nombre",            f'=COUNTA({DB}!C3:C5000)',                                                       "tienen nombre registrado"),
    ("Solo teléfono",         f'=COUNTIFS({DB}!C3:C5000,"",{DB}!D3:D5000,"<>")',                               "sin nombre pero con tel"),
    ("Inscritos",             f'=COUNTIF({DB}!K3:K5000,"Inscrito")',                                           "resultado = Inscrito"),
    ("En proceso inscripción",f'=COUNTIF({DB}!K3:K5000,"En Proceso de Inscripción")',                          "casi convertidos"),
    ("Interesados",           f'=COUNTIF({DB}!K3:K5000,"Interesado")',                                         "seguimiento activo"),
    ("Desinteresados",        f'=COUNTIF({DB}!K3:K5000,"Desinteresado")',                                      "no convertidos"),
    ("Sin Respuesta",         f'=COUNTIF({DB}!K3:K5000,"Sin Respuesta")',                                      "no contestaron"),
    ("Tasa Conversión",       f'=IFERROR(COUNTIF({DB}!K3:K5000,"Inscrito")/COUNTA({DB}!C3:C5000),0)',          "sobre leads con nombre"),
]

for i, (label, formula, detail) in enumerate(KPI_ROWS, 3):
    pv_label(ws_pv, i, 1, label, C_LGRAY if i%2==0 else C_WHITE)
    fmt = "0.0%" if "Tasa" in label else "#,##0"
    bg_val = C_GREEN if label=="Inscritos" else (C_ACC2 if label=="Desinteresados" else C_WHITE)
    pv_cell(ws_pv, i, 2, formula, fmt=fmt, bg=bg_val, bold=(label in ("Inscritos","Tasa Conversión")))
    pv_cell(ws_pv, i, 3, detail)
    pv_cell(ws_pv, i, 4, "")


# ── Sección B: Por Resultado (col 6) ──────────────────────────────────────
pv_title(ws_pv, 1, 6, "📈  POR RESULTADO", 4, C_ACC2)
for ci, h in [(6,"RESULTADO"),(7,"TOTAL"),(8,"% LEADS"),(9,"")]:
    pv_hdr(ws_pv, 2, ci, h, C_ACC2)

for i, res in enumerate(LISTA_RESULTADO, 3):
    pv_label(ws_pv, i, 6, res, C_LGRAY if i%2==0 else C_WHITE)
    pv_cell(ws_pv,  i, 7, f'=COUNTIF({DB}!K3:K5000,"{res}")')
    pv_cell(ws_pv,  i, 8, f'=IFERROR(G{i}/COUNTA({DB}!C3:C5000),0)', fmt="0.0%")
    pv_cell(ws_pv,  i, 9, "")

# ── Sección C: Por Fuente (col 11) ────────────────────────────────────────
pv_title(ws_pv, 1, 11, "📡  POR FUENTE", 4, C_MID)
for ci, h in [(11,"FUENTE"),(12,"LEADS"),(13,"INSCRITOS"),(14,"TASA CONV.")]:
    pv_hdr(ws_pv, 2, ci, h, C_MID)

for i, fuente in enumerate(LISTA_FUENTE, 3):
    pv_label(ws_pv, i, 11, fuente, C_LGRAY if i%2==0 else C_WHITE)
    pv_cell(ws_pv,  i, 12, f'=COUNTIF({DB}!H3:H5000,"{fuente}")')
    pv_cell(ws_pv,  i, 13, f'=COUNTIFS({DB}!H3:H5000,"{fuente}",{DB}!K3:K5000,"Inscrito")')
    pv_cell(ws_pv,  i, 14, f'=IFERROR(M{i}/L{i},0)', fmt="0.0%")


# ── Sección D: Por Interés — filas dinámicas (col 1, desde fila 16) ───────
START_INT = 16
pv_title(ws_pv, START_INT, 1, "🎯  POR INTERÉS", 4, C_ACC1)
for ci, h in [(1,"INTERÉS"),(2,"LEADS"),(3,"INSCRITOS"),(4,"TASA")]:
    pv_hdr(ws_pv, START_INT+1, ci, h, C_ACC1)

# Usamos COUNTIF con ISNUMBER(SEARCH()) para capturar valores múltiples
# ej. "Academia de Fútbol / Curso de Verano" se cuenta en ambos
for i, interes in enumerate(LISTA_INTERESES, START_INT+2):
    safe = interes.replace('"', "'")
    pv_label(ws_pv, i, 1, interes, C_LGRAY if i%2==0 else C_WHITE)
    pv_cell(ws_pv,  i, 2, f'=COUNTIF({DB}!F3:F5000,"*{safe}*")')
    pv_cell(ws_pv,  i, 3, f'=COUNTIFS({DB}!F3:F5000,"*{safe}*",{DB}!K3:K5000,"Inscrito")')
    pv_cell(ws_pv,  i, 4, f'=IFERROR(C{i}/B{i},0)', fmt="0.0%")

# ── Sección E: Por Mes (col 6, desde fila 16) ─────────────────────────────
pv_title(ws_pv, START_INT, 6, "📅  EVOLUCIÓN MENSUAL", 4, C_MID)
for ci, h in [(6,"MES"),(7,"LEADS"),(8,"INSCRITOS"),(9,"TASA")]:
    pv_hdr(ws_pv, START_INT+1, ci, h, C_MID)

meses_reales = sorted(set(l["mes"] for l in leads if l["mes"]))
for i, mes in enumerate(meses_reales, START_INT+2):
    pv_label(ws_pv, i, 6, mes, C_LGRAY if i%2==0 else C_WHITE)
    pv_cell(ws_pv,  i, 7, f'=COUNTIF({DB}!N3:N5000,"{mes}")')
    pv_cell(ws_pv,  i, 8, f'=COUNTIFS({DB}!N3:N5000,"{mes}",{DB}!K3:K5000,"Inscrito")')
    pv_cell(ws_pv,  i, 9, f'=IFERROR(H{i}/G{i},0)', fmt="0.0%")

# ── Sección F: Por Semana (col 11, desde fila 16) ─────────────────────────
pv_title(ws_pv, START_INT, 11, "📆  EVOLUCIÓN SEMANAL", 4, C_ACC2)
for ci, h in [(11,"SEMANA"),(12,"LEADS"),(13,"INSCRITOS"),(14,"TASA")]:
    pv_hdr(ws_pv, START_INT+1, ci, h, C_ACC2)

semanas_reales = sorted(set(l["semana"] for l in leads if l["semana"]))
for i, sem in enumerate(semanas_reales, START_INT+2):
    pv_label(ws_pv, i, 11, sem, C_LGRAY if i%2==0 else C_WHITE)
    pv_cell(ws_pv,  i, 12, f'=COUNTIF({DB}!M3:M5000,"{sem}")')
    pv_cell(ws_pv,  i, 13, f'=COUNTIFS({DB}!M3:M5000,"{sem}",{DB}!K3:K5000,"Inscrito")')
    pv_cell(ws_pv,  i, 14, f'=IFERROR(M{i}/L{i},0)', fmt="0.0%")

# ── Sección G: Por Medio (col 1, después de intereses) ────────────────────
START_MED = START_INT + len(LISTA_INTERESES) + 4
pv_title(ws_pv, START_MED, 1, "📱  POR MEDIO DE CONTACTO", 4, C_ACC2)
for ci, h in [(1,"MEDIO"),(2,"LEADS"),(3,"INSCRITOS"),(4,"TASA")]:
    pv_hdr(ws_pv, START_MED+1, ci, h, C_ACC2)

for i, medio in enumerate(LISTA_MEDIO, START_MED+2):
    pv_label(ws_pv, i, 1, medio, C_LGRAY if i%2==0 else C_WHITE)
    pv_cell(ws_pv,  i, 2, f'=COUNTIF({DB}!G3:G5000,"{medio}")')
    pv_cell(ws_pv,  i, 3, f'=COUNTIFS({DB}!G3:G5000,"{medio}",{DB}!K3:K5000,"Inscrito")')
    pv_cell(ws_pv,  i, 4, f'=IFERROR(C{i}/B{i},0)', fmt="0.0%")

print(f"   ✓ DATOS_PIVOT (secciones: KPIs, Resultado, Fuente, Interés, Mes, Semana, Medio)")


# ══════════════════════════════════════════════════════════════════════════
# HOJA: DASHBOARD
# ══════════════════════════════════════════════════════════════════════════
print("▶ Creando DASHBOARD…")
ws_d = wb.create_sheet("DASHBOARD")
ws_d.sheet_properties.tabColor = "E94560"
ws_d.sheet_view.showGridLines = False

PV = "DATOS_PIVOT"

# Anchos columnas A-T
col_widths = [1,14,14,14,14,14,14,2,18,18,18,18,18,2,14,14,14,14,14,2]
for ci, w in enumerate(col_widths, 1):
    ws_d.column_dimensions[get_column_letter(ci)].width = w

# Alturas de filas
for r, h in [(1,45),(2,18),(3,10),(4,12),(5,40),(6,40),(7,14),(8,16),(9,16),(10,16),(11,16),(12,16),(13,16),(14,16),(15,14)]:
    ws_d.row_dimensions[r].height = h

def dash_bg(ws, r1, c1, r2, c2, hex_c):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(r, c).fill = fill(hex_c)

dash_bg(ws_d, 1, 1, 80, 20, C_DARK)

# ── Título ────────────────────────────────────────────────────────────────
ws_d.merge_cells("B1:G1")
t = ws_d["B1"]
t.value = "DASHBOARD COMERCIAL"
t.font  = fnt(bold=True, sz=22, color=C_WHITE)
t.alignment = aln()

ws_d.merge_cells("I1:T1")
t2 = ws_d["I1"]
t2.value = "Andres Chitiva  ·  Julio 2026"
t2.font  = fnt(sz=13, color=C_GRAY, italic=True)
t2.alignment = aln(h="right")

ws_d.merge_cells("B2:T2")
sub = ws_d["B2"]
sub.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Las métricas se actualizan automáticamente al editar LEADS_DB"
sub.font  = fnt(sz=8, color=C_GRAY, italic=True)
sub.alignment = aln(h="left")

# ── Tarjetas KPI (fila 4-6) ───────────────────────────────────────────────
# Fórmulas apuntan a DATOS_PIVOT fila 3-11
# KPI cards: Total, Inscritos, Desinteresados, Sin Resp, Interesados, Tasa
KPI_CARDS = [
    ("B","C", C_ACC1,   "TOTAL LEADS",      f"=DATOS_PIVOT!B3",  "leads registrados"),
    ("D","E", C_GREEN,  "INSCRITOS",         f"=DATOS_PIVOT!B6",  f"de {total} leads"),
    ("F","G", C_ACC2,   "DESINTERESADOS",    f"=DATOS_PIVOT!B9",  "no convertidos"),
    ("I","J", C_YELLOW, "INTERESADOS",       f"=DATOS_PIVOT!B8",  "seguimiento activo"),
    ("K","L", C_ORANGE, "SIN RESPUESTA",     f"=DATOS_PIVOT!B10", "no contestaron"),
    ("M","N", C_BLUE,   "TASA CONVERSIÓN",   f"=DATOS_PIVOT!B11", "sobre leads con nombre"),
]

for (c1l, c2l, bg, label, formula, sub_txt) in KPI_CARDS:
    c1 = column_index_from_string(c1l)
    c2 = column_index_from_string(c2l)
    for r in (4,5,6):
        ws_d.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        for col in range(c1, c2+1):
            ws_d.cell(r, col).fill = fill(bg)
    lbl = ws_d.cell(4, c1, label)
    lbl.font = fnt(bold=True, sz=8, color=C_TEXT if bg==C_YELLOW else C_WHITE)
    lbl.alignment = aln()
    val = ws_d.cell(5, c1, formula)
    is_pct = "TASA" in label
    val.number_format = "0.0%" if is_pct else "#,##0"
    val.font = fnt(bold=True, sz=24, color=C_TEXT if bg==C_YELLOW else C_WHITE)
    val.alignment = aln()
    sbt = ws_d.cell(6, c1, sub_txt)
    sbt.font = fnt(sz=8, color=C_TEXT if bg==C_YELLOW else "DFE6E9", italic=True)
    sbt.alignment = aln()


# ── Tabla Resultado en Dashboard (filas 8-16, cols B-G) ───────────────────
def dash_section(ws, row, c1, c2, text, bg=C_ACC1):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row, c1, text)
    c.font = fnt(bold=True, sz=10, color=C_WHITE)
    c.fill = fill(bg)
    c.alignment = aln(h="left")

def dash_hdr_row(ws, row, cols_hdrs, bg=C_MID):
    for col, h in cols_hdrs:
        c = ws.cell(row, col, h)
        c.font = fnt(bold=True, sz=8, color=C_WHITE)
        c.fill = fill(bg)
        c.alignment = aln()
        c.border = bdr(bg)

def dash_row(ws, row, col_vals, alt=False):
    bg_r = C_LGRAY if alt else "ECEFF1"
    for col, val, fmt in col_vals:
        c = ws.cell(row, col, val)
        c.font = fnt(sz=9)
        c.fill = fill(bg_r)
        c.alignment = aln()
        c.border = bdr()
        if fmt: c.number_format = fmt

# RESULTADO
dash_section(ws_d, 7, 2, 7, "📊  DISTRIBUCIÓN POR RESULTADO", C_ACC2)
dash_hdr_row(ws_d, 8, [(2,"RESULTADO"),(5,"TOTAL"),(6,"%"),(7,"OBS")], C_ACC2)
ws_d.merge_cells("B8:D8")

OBS_MAP = {
    "Inscrito":                  "✅ Convertido",
    "En Proceso de Inscripción": "⏳ Cerrar pronto",
    "Clase Muestra":             "🏃 Mostrar valor",
    "Contactado/Activo":         "👀 En seguimiento",
    "Interesado":                "🔥 Trabajar ahora",
    "En Espera":                 "⏸ Esperar respuesta",
    "Desinteresado":             "❌ No convertido",
    "Sin Respuesta":             "📵 Reactivar",
}
# Row offset in DATOS_PIVOT for Resultado: starts at row 3 (col G=7)
PV_RES_START = 3
for i, res in enumerate(LISTA_RESULTADO, 9):
    pv_row = PV_RES_START + i - 9   # fila en DATOS_PIVOT sección resultado (col G)
    ws_d.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
    alt = i % 2 == 0
    bg_r = RESULTADO_COLOR.get(res, (C_LGRAY, C_TEXT))[0]
    fg_r = RESULTADO_COLOR.get(res, (C_LGRAY, C_TEXT))[1]
    for col in (2,5,6,7):
        ws_d.cell(i, col).fill = fill(C_LGRAY if alt else C_WHITE)
        ws_d.cell(i, col).border = bdr()
        ws_d.cell(i, col).alignment = aln()
        ws_d.cell(i, col).font = fnt(sz=9)
    lc = ws_d.cell(i, 2, res)
    lc.fill = fill(bg_r); lc.font = fnt(sz=9, bold=True, color=fg_r)
    # Formula apunta a columna G de DATOS_PIVOT (sección resultado)
    ws_d.cell(i, 5, f"=DATOS_PIVOT!G{pv_row}").number_format = "#,##0"
    ws_d.cell(i, 6, f"=DATOS_PIVOT!H{pv_row}").number_format = "0.0%"
    ws_d.cell(i, 7, OBS_MAP.get(res, ""))

# FUENTE  (cols I-N, filas 7-18)
dash_section(ws_d, 7, 9, 14, "📡  LEADS POR FUENTE", C_MID)
dash_hdr_row(ws_d, 8, [(9,"FUENTE"),(11,"LEADS"),(12,"INSCRITOS"),(13,"TASA"),(14,"% TOTAL")], C_MID)
ws_d.merge_cells("I8:J8")

PV_FUE_START = 3
for i, fuente in enumerate(LISTA_FUENTE, 9):
    pv_row = PV_FUE_START + i - 9
    ws_d.merge_cells(start_row=i, start_column=9, end_row=i, end_column=10)
    for col in (9,11,12,13,14):
        ws_d.cell(i,col).fill = fill(C_LGRAY if i%2==0 else C_WHITE)
        ws_d.cell(i,col).border = bdr()
        ws_d.cell(i,col).alignment = aln()
        ws_d.cell(i,col).font = fnt(sz=9)
    ws_d.cell(i, 9, fuente)
    ws_d.cell(i,11, f"=DATOS_PIVOT!L{pv_row}").number_format = "#,##0"
    ws_d.cell(i,12, f"=DATOS_PIVOT!M{pv_row}").number_format = "#,##0"
    ws_d.cell(i,13, f"=DATOS_PIVOT!N{pv_row}").number_format = "0.0%"
    ws_d.cell(i,14, f"=IFERROR(DATOS_PIVOT!L{pv_row}/DATOS_PIVOT!B3,0)").number_format = "0.0%"


# ── Evolución Mensual (filas 17-xx, cols B-G) ─────────────────────────────
EVO_START = 17
dash_section(ws_d, EVO_START,   2,  7, "📅  EVOLUCIÓN MENSUAL",  C_ACC1)
dash_section(ws_d, EVO_START,   9, 14, "📆  EVOLUCIÓN SEMANAL",  C_ACC2)
dash_hdr_row(ws_d, EVO_START+1, [(2,"MES"),(4,"LEADS"),(5,"INSCRITOS"),(6,"TASA"),(7,"VAR")], C_ACC1)
ws_d.merge_cells(f"B{EVO_START+1}:C{EVO_START+1}")
dash_hdr_row(ws_d, EVO_START+1, [(9,"SEMANA"),(11,"LEADS"),(12,"INSCRITOS"),(13,"TASA")], C_ACC2)
ws_d.merge_cells(f"I{EVO_START+1}:J{EVO_START+1}")

for i, mes in enumerate(meses_reales, EVO_START+2):
    pv_mes_row = (START_INT+2) + i - (EVO_START+2)
    ws_d.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
    alt = i%2==0
    for col in (2,4,5,6,7):
        ws_d.cell(i,col).fill = fill(C_LGRAY if alt else "ECEFF1")
        ws_d.cell(i,col).border = bdr()
        ws_d.cell(i,col).alignment = aln()
        ws_d.cell(i,col).font = fnt(sz=9)
    ws_d.cell(i,2, mes)
    ws_d.cell(i,4, f"=DATOS_PIVOT!G{pv_mes_row}").number_format = "#,##0"
    ws_d.cell(i,5, f"=DATOS_PIVOT!H{pv_mes_row}").number_format = "#,##0"
    ws_d.cell(i,6, f"=DATOS_PIVOT!I{pv_mes_row}").number_format = "0.0%"
    pv_prev = pv_mes_row - 1
    ws_d.cell(i,7, f"=IFERROR((DATOS_PIVOT!G{pv_mes_row}-DATOS_PIVOT!G{pv_prev})/DATOS_PIVOT!G{pv_prev},0)").number_format = "+0.0%;-0.0%;0.0%"
    ws_d.cell(i,7).font = fnt(sz=9, color=C_GREEN)

for i, sem in enumerate(semanas_reales, EVO_START+2):
    pv_sem_row = (START_INT+2) + i - (EVO_START+2)
    ws_d.merge_cells(start_row=i, start_column=9, end_row=i, end_column=10)
    alt = i%2==0
    for col in (9,11,12,13):
        ws_d.cell(i,col).fill = fill(C_LGRAY if alt else "ECEFF1")
        ws_d.cell(i,col).border = bdr()
        ws_d.cell(i,col).alignment = aln()
        ws_d.cell(i,col).font = fnt(sz=9)
    ws_d.cell(i, 9, sem)
    ws_d.cell(i,11, f"=DATOS_PIVOT!L{pv_sem_row}").number_format = "#,##0"
    ws_d.cell(i,12, f"=DATOS_PIVOT!M{pv_sem_row}").number_format = "#,##0"
    ws_d.cell(i,13, f"=DATOS_PIVOT!N{pv_sem_row}").number_format = "0.0%"


# ── TOP INTERÉS (debajo de evo mensual, cols B-G) ─────────────────────────
INT_START = EVO_START + max(len(meses_reales), len(semanas_reales)) + 4
dash_section(ws_d, INT_START, 2, 8, "🎯  TOP INTERESES", C_MID)
dash_hdr_row(ws_d, INT_START+1, [(2,"INTERÉS"),(5,"LEADS"),(6,"INSCRITOS"),(7,"TASA"),(8,"% TOTAL")], C_MID)
ws_d.merge_cells(f"B{INT_START+1}:D{INT_START+1}")

# Contamos desde los datos reales para ordenar por volumen
interes_count = Counter()
for l in leads:
    for part in l["interes"].split(" / "):
        p = part.strip()
        if p: interes_count[p] += 1
top_intereses = [k for k,_ in interes_count.most_common()]

for i, interes in enumerate(top_intereses[:15], INT_START+2):
    pv_int_row = (START_INT+2) + LISTA_INTERESES.index(interes) if interes in LISTA_INTERESES else None
    ws_d.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
    alt = i%2==0
    for col in (2,5,6,7,8):
        ws_d.cell(i,col).fill = fill(C_LGRAY if alt else "ECEFF1")
        ws_d.cell(i,col).border = bdr()
        ws_d.cell(i,col).alignment = aln()
        ws_d.cell(i,col).font = fnt(sz=9)
    ws_d.cell(i, 2, interes)
    if pv_int_row:
        ws_d.cell(i,5, f"=DATOS_PIVOT!B{pv_int_row}").number_format = "#,##0"
        ws_d.cell(i,6, f"=DATOS_PIVOT!C{pv_int_row}").number_format = "#,##0"
        ws_d.cell(i,7, f"=DATOS_PIVOT!D{pv_int_row}").number_format = "0.0%"
        ws_d.cell(i,8, f"=IFERROR(DATOS_PIVOT!B{pv_int_row}/DATOS_PIVOT!B3,0)").number_format = "0.0%"
    else:
        safe = interes.replace('"',"'")
        ws_d.cell(i,5, f'=COUNTIF(LEADS_DB!F3:F5000,"*{safe}*")').number_format = "#,##0"
        ws_d.cell(i,6, f'=COUNTIFS(LEADS_DB!F3:F5000,"*{safe}*",LEADS_DB!K3:K5000,"Inscrito")').number_format = "#,##0"
        ws_d.cell(i,7, f"=IFERROR(F{i}/E{i},0)").number_format = "0.0%"
        ws_d.cell(i,8, f"=IFERROR(E{i}/DATOS_PIVOT!B3,0)").number_format = "0.0%"

print("   ✓ DASHBOARD (KPIs, tablas resultado/fuente/evo/intereses)")


# ══════════════════════════════════════════════════════════════════════════
# GRÁFICOS — vinculados a DATOS_PIVOT (se actualizan solos)
# ══════════════════════════════════════════════════════════════════════════
print("▶ Creando gráficos…")

# Hoja oculta con datos de gráficos, calculados por fórmulas
ws_ch = wb.create_sheet("_GrafData")
ws_ch.sheet_state = "hidden"

# Datos resultado (para dona/barras) — copia de DATOS_PIVOT col G-H
ws_ch["A1"] = "Resultado"; ws_ch["B1"] = "Total"
for i, res in enumerate(LISTA_RESULTADO, 2):
    ws_ch.cell(i, 1, res)
    ws_ch.cell(i, 2, f"=DATOS_PIVOT!G{i}")   # fila i en sección resultado de DATOS_PIVOT

# Datos fuente — copia de DATOS_PIVOT col L-N
ws_ch["D1"] = "Fuente"; ws_ch["E1"] = "Leads"; ws_ch["F1"] = "Inscritos"
for i, fuente in enumerate(LISTA_FUENTE, 2):
    ws_ch.cell(i, 4, fuente)
    ws_ch.cell(i, 5, f"=DATOS_PIVOT!L{i}")
    ws_ch.cell(i, 6, f"=DATOS_PIVOT!M{i}")

# Datos mensuales — copia de DATOS_PIVOT sección mes
ws_ch["H1"] = "Mes"; ws_ch["I1"] = "Leads"; ws_ch["J1"] = "Inscritos"
for i, mes in enumerate(meses_reales, 2):
    pv_row = (START_INT+2) + i - 2
    ws_ch.cell(i, 8, mes)
    ws_ch.cell(i, 9, f"=DATOS_PIVOT!G{pv_row}")
    ws_ch.cell(i,10, f"=DATOS_PIVOT!H{pv_row}")

n_res = len(LISTA_RESULTADO)
n_fue = len(LISTA_FUENTE)
n_mes = len(meses_reales)

# ── Gráfico 1: Barras horizontales — Leads por Resultado ─────────────────
chart1 = BarChart()
chart1.type = "bar"
chart1.title = "Leads por Resultado"
chart1.style = 10
chart1.width = 16; chart1.height = 10
chart1.grouping = "clustered"
chart1.y_axis.title = "Resultado"
chart1.x_axis.title = "Leads"
data1 = Reference(ws_ch, min_col=2, max_col=2, min_row=1, max_row=n_res+1)
cats1 = Reference(ws_ch, min_col=1, min_row=2, max_row=n_res+1)
chart1.add_data(data1, titles_from_data=True)
chart1.set_categories(cats1)
chart1.series[0].graphicalProperties.solidFill = C_ACC1
ws_d.add_chart(chart1, "B22")

# ── Gráfico 2: Barras agrupadas — Leads e Inscritos por Fuente ───────────
chart2 = BarChart()
chart2.type = "bar"
chart2.title = "Leads por Fuente"
chart2.style = 10
chart2.width = 16; chart2.height = 10
chart2.grouping = "clustered"
data2 = Reference(ws_ch, min_col=5, max_col=6, min_row=1, max_row=n_fue+1)
cats2 = Reference(ws_ch, min_col=4, min_row=2, max_row=n_fue+1)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.series[0].graphicalProperties.solidFill = C_MID
chart2.series[1].graphicalProperties.solidFill = C_GREEN
ws_d.add_chart(chart2, "J22")

# ── Gráfico 3: Líneas — Evolución Mensual ────────────────────────────────
chart3 = LineChart()
chart3.title = "Evolución Mensual"
chart3.style = 10
chart3.width = 22; chart3.height = 12
data3 = Reference(ws_ch, min_col=9, max_col=10, min_row=1, max_row=n_mes+1)
cats3 = Reference(ws_ch, min_col=8, min_row=2, max_row=n_mes+1)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
chart3.series[0].graphicalProperties.line.solidFill = C_ACC1
chart3.series[0].graphicalProperties.line.width = 25000
chart3.series[1].graphicalProperties.line.solidFill = C_GREEN
chart3.series[1].graphicalProperties.line.width = 25000
ws_d.add_chart(chart3, "B38")

print("   ✓ Gráficos (resultado, fuente, evolución mensual)")


# ══════════════════════════════════════════════════════════════════════════
# REORDENAR HOJAS y GUARDAR
# ══════════════════════════════════════════════════════════════════════════
ORDER = ["LEADS_DB", "DATOS_PIVOT", "DASHBOARD", "Config", "_GrafData"]
wb._sheets = sorted(wb._sheets, key=lambda s: ORDER.index(s.title) if s.title in ORDER else 99)

wb.save(DST)
print(f"\n✅  Guardado: {DST}")
print(f"   Hojas:   {[s.title for s in wb.worksheets]}")
print(f"   Leads:   {total} filas útiles  (con nombre: {con_nombre}, solo tel: {sin_nombre})")
